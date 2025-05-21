import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import json
from io import StringIO
import base64
import os
from datetime import datetime
from libs.Helpers import log_error_wr


def encode_string(value: str) -> str:
    encoded_bytes = base64.b64encode(value.encode())
    return encoded_bytes.decode()

def decode_string(value: str) -> str:
    decoded_bytes = base64.b64decode(value.encode())
    return decoded_bytes.decode()

def encode_file_to_base64(filepath):
    with open(filepath, "rb") as file:
        encoded_string = base64.b64encode(file.read()).decode("utf-8")
    return encoded_string


def save_xlsx_locally(base64_pdf_string, filename):
    try:
        images_folder = 'flask_files'
        wr_folder = "warehouse"
        excel_folder = "excel_files"

        desktop_path = os.path.expanduser("~/Desktop")
        folder_path = os.path.join(desktop_path, images_folder, wr_folder, excel_folder)
        os.makedirs(folder_path, exist_ok=True)
        if len(base64_pdf_string.split(',')) > 1:
            _, base64_string = base64_pdf_string.split(',', 1)
        else:
            base64_string = base64_pdf_string
        file_data = base64.b64decode(base64_string)
        file_path = os.path.join(folder_path, filename)
        with open(file_path, 'wb') as f:
            f.write(file_data)
        return True
        
    except Exception as error:
        log_error_wr(f"{error} ---------- {base64_pdf_string}")
        return False

def get_xlsx_locally(filename):
    try:
        images_folder = 'flask_files'
        wr_folder = "warehouse"
        excel_folder = "excel_files"

        filename = filename.replace(".xlsx","")

        desktop_path = os.path.expanduser("~/Desktop")
        folder_path = os.path.join(desktop_path, images_folder, wr_folder, excel_folder)
        os.makedirs(folder_path, exist_ok=True)

        file_path = os.path.join(folder_path, f"{filename}.xlsx")
        with open(file_path, 'rb') as f:
            pdf_content = f.read()
            pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
            return pdf_base64
        
    except Exception as error:
        log_error_wr(error)
        return False


def convert_data_to_XLSX(json_data, main_header):
    try:
        now = datetime.now()
        year = now.year
        month = str(now.month).zfill(2)
        day = str(now.day).zfill(2)
        hours = str(now.hour).zfill(2)
        minutes = str(now.minute).zfill(2)

        temp_filename = f"local_storage/Stok-{year}-{month}-{day}-{hours}-{minutes}.xlsx"

        json_string = json.dumps(json_data)
        json_data_io = StringIO(json_string)

        data = pd.read_json(json_data_io)
        data.to_excel(temp_filename, index=False)
        workbook = load_workbook(temp_filename)
        sheet = workbook.active

        sheet.insert_rows(1)
        tepe_baslik = main_header
        baslik_hucre = sheet.cell(row=1, column=1)
        baslik_hucre.value = tepe_baslik

        max_column = sheet.max_column
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)

        baslik_hucre.alignment = Alignment(horizontal="center")
        baslik_hucre.font = Font(size=14, bold=True)

        # Ortalama ve Hizalamalar
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        workbook.save(temp_filename)

        xlsx_b64_data = encode_file_to_base64(temp_filename)

        save_xlsx_locally(xlsx_b64_data, temp_filename)

        final_data = {
            "status" : "ok",
            "b64_data" : xlsx_b64_data
        }
    except Exception as error:
        print(error)
        final_data = {
            "status" : "error",
            "error_text" : str(error)
        }
    
    finally:
        return final_data